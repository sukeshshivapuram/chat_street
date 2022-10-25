# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.



import io
import locale
import base64
import textwrap
from copy import copy

from datetime import datetime
from openpyxl import Workbook
from odoo import models, fields, api,_
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from odoo.exceptions import UserError, ValidationError,Warning
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
import pdb

class StockAdjustment(models.TransientModel):
    _name = "stock.adjustment.report"
    _description = "stock.adjustment.report"

    date_start = fields.Datetime(string="Start Date", required=True)
    date_end = fields.Datetime(string="End Date", required=True, default=fields.Datetime.now)
    sa_report = fields.Binary('Astock Adjustment Report')
    file_name = fields.Char('File Name')
    # s_report_printed = fields.Boolean('MR Printed')
    sa_printed = fields.Boolean('Astock Adjustment Printed')
    # company_id = fields.Many2one('res.company',string='Company')
    # warehouse_ids = fields.Many2many('stock.warehouse',string='Warehouse',required="1")
    # location_id = fields.Many2one('stock.location',string='Location', domain="[('usage','!=','view')]")
    # production_id = fields.Many2one('stock.move',string='Product Id')
    start_date = fields.Date('Start Date')
    end_date = fields.Date('End Date')
    # grn_type = fields.Selection([('internal','Internal'),('external','External')],string='GRN Type')
    # category_id = fields.Many2many('product.category',string='Category')
    # product_ids = fields.Many2many('product.product',string='Products')


    @api.constrains('date_start')
    def _code_constrains(self):
        if self.date_start > self.date_end:
            raise ValidationError(_("'Start Date' must be before 'End Date'"))


    
    def generate_report(self):

    
        # #Create Workbook and Worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Adjustment"
        report_heading = " Stock Adjustment from" + ' ' + datetime.strftime(self.date_start, '%d-%m-%Y') + ' '+ 'To' + ' '+ datetime.strftime( self.date_end, '%d-%m-%Y')
        
        # #Border
        thin = Side(border_style="thin", color="000000")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=18)

        comapny = ws.cell(row=1, column=1, value=self.env.user.company_id.name + report_heading)
        comapny.alignment = Alignment(horizontal='center',vertical='center')
        comapny.border = Border(bottom=thin,top=thin)
        comapny.font = Font(size=10,name='Calibri')


        # Defining the Table Caolumn headings
        sl_no = ws.cell(row=3, column=1, value="S.No")
        sl_no.alignment = copy(comapny.alignment)
        warehouse_id = ws.cell(row=3, column=2, value="Site")
        warehouse_id.alignment = copy(comapny.alignment)

        org_state = ws.cell(row=3, column=3, value="Orig. State")
        org_state.alignment = copy(comapny.alignment)

        final_state = ws.cell(row=3, column=4, value="Final State")
        final_state.alignment = Alignment(horizontal='center',vertical='center')

        date = ws.cell(row=3, column=5, value="Date")
        date.alignment = Alignment(horizontal='center',vertical='center')

        created_by = ws.cell(row=3, column=6, value="Create By")
        created_by.alignment = Alignment(horizontal='center',vertical='center')
        
        created_on = ws.cell(row=3, column=7, value="Create On")
        created_on.alignment = Alignment(horizontal='center',vertical='center')

        product_id = ws.cell(row=3, column=8, value="Item")
        product_id.alignment = Alignment(horizontal='center',vertical='center')

        qty = ws.cell(row=3, column=9, value="Qty")
        qty.alignment = Alignment(horizontal='center',vertical='center')

        product_unit = ws.cell(row=3, column=10, value="Unit")
        product_unit.alignment = Alignment(horizontal='center',vertical='center')
        
        amount = ws.cell(row=3, column=11, value="Amount")
        amount.alignment = Alignment(horizontal='center',vertical='center')


        query='''

        select 
        sm.date::date as sm_date, 
        sm.create_date::date as sm_createdate,
        sm.create_uid ,
        sm.price_unit ,
        sm.inventory_id ,
        sm.product_uom_qty,
        sm.product_id,
        sm.location_id ,
        sm.loss_reason as loss,
        sm.location_dest_id 

        from stock_move as sm

        where  
        (sm.location_id =14 or 
        sm.location_dest_id =14) and 
        sm.state ='done' and 
        sm.date > %s and 
        sm.date < %s 
        '''

        params = (self.date_start ,self.date_end)

        self.env.cr.execute(query, params)
        result = self.env.cr.dictfetchall()

        

  
        row_c=4
        sl_num=1

        for stock_line in result:
            if stock_line['location_id'] != 14:
                c_name = self.env['stock.location'].search([('id','=',stock_line['location_id'])])
            elif stock_line['location_dest_id'] != 14:
                c_name = self.env['stock.location'].search([('id','=',stock_line['location_dest_id'])])
            if c_name.complete_name:
                w_code=c_name.complete_name.split('/')

                c_name = self.env['stock.warehouse'].search([('code','=',w_code[0])])
            p_quer='''
            select 
                pt.name as p_name,
                uu.name as uu_name,
                pc.complete_name  as pc_name
            from  
                product_product as pp,
                product_template as pt,
                uom_uom as uu,
                product_category as pc
            where 
                product_tmpl_id = pt.id and 
                pp.id =%s and
                uu.id = pt.uom_id and
                pc.id = pt.categ_id
            '''
            self.env.cr.execute(p_quer, [stock_line['product_id']])
            product = self.env.cr.dictfetchall()


            sl_val = ws.cell(row=row_c, column=1, value=sl_num)
            sl_val.alignment = Alignment(horizontal='center',vertical='center')

            warehouse_id = ws.cell(row=row_c, column=2, value=c_name.name)

            org_state = ws.cell(row=row_c, column=3, value="Good" )
            org_state .alignment = Alignment(horizontal='center',vertical='center')

            final_state = ws.cell(row=row_c, column=4, value=stock_line['loss'])
            final_state .alignment = Alignment(horizontal='center',vertical='center')

            date = ws.cell(row=row_c, column=5, value=stock_line['sm_date'])
            date.alignment = Alignment(horizontal='center',vertical='center')
            # if stock_line['create_uid']:
            c_name = self.env['res.users'].search([('id','=',stock_line['create_uid'])])
            if c_name:
                users_name = c_name.name
            else:
                users_name ='Administrator'
            

            created_by = ws.cell(row=row_c, column=6, value=users_name)
            created_by.alignment = Alignment(horizontal='center',vertical='center')

            
            created_on = ws.cell(row=row_c, column=7, value=stock_line['sm_createdate'])
            created_on .alignment = Alignment(horizontal='center',vertical='center')
            if stock_line['product_id']:
                p_name = self.env['product.product'].search([('id','=',stock_line['product_id'])])
            
            
            product_id = ws.cell(row=row_c, column=8, value=product[0]['p_name'])
            # product_id.alignment = Alignment(horizontal='center',vertical='center')
            real_qty = 0
            if stock_line['location_dest_id'] == 14:
                real_qty = -stock_line['product_uom_qty']
            else:
                real_qty = stock_line['product_uom_qty']

            qty = ws.cell(row=row_c, column=9, value=real_qty)
            # qty.alignment = Alignment(horizontal='center',vertical='center')
            if not stock_line['price_unit']:
                p_unit = p_name.standard_price
            else:
                p_unit =stock_line['price_unit']
            product_unit = ws.cell(row=row_c, column=10, value= p_unit)

            amount = ws.cell(row=row_c, column=11, value=real_qty *p_unit)
            

            sl_num +=1
            row_c +=1



       




        fp = io.BytesIO()
        wb.save(fp)
        excel_file = base64.encodestring(fp.getvalue())
        self.sa_report = excel_file
        self.sa_printed = True
        self.file_name = "Stock Adjustment.xlsx"
        fp.close()

        return {
        'view_mode': 'form',
        'res_id': self.id,
        'res_model': 'stock.adjustment.report',
        'view_type': 'form',
        'type': 'ir.actions.act_window',
        'context': self.env.context,
        'target': 'new',
                   }


   
