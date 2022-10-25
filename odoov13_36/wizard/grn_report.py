# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.



import io
import locale
import base64
import textwrap
from copy import copy

from datetime import datetime
from openpyxl import Workbook
from odoo import models, fields, api,_
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from odoo.exceptions import UserError, ValidationError,Warning
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
import pdb

class GRNReport(models.TransientModel):
    _name = "grn.report"
    _description = "mr Report"

    date_start = fields.Datetime(string="Start Date", required=True)
    date_end = fields.Datetime(string="End Date", required=True, default=fields.Datetime.now)
    grn_report = fields.Binary('GRN Report')
    file_name = fields.Char('File Name')
    # s_report_printed = fields.Boolean('MR Printed')
    grn_printed = fields.Boolean('GRN Printed')
    # company_id = fields.Many2one('res.company',string='Company')
    # warehouse_ids = fields.Many2many('stock.warehouse',string='Warehouse',required="1")
    # location_id = fields.Many2one('stock.location',string='Location', domain="[('usage','!=','view')]")
    # production_id = fields.Many2one('stock.move',string='Product Id')
    start_date = fields.Date('Start Date')
    end_date = fields.Date('End Date')
    grn_type = fields.Selection([('internal','Internal'),('external','External')],string='GRN Type',required=True)
    category_id = fields.Many2many('product.category',string='Category')
    product_ids = fields.Many2many('product.product',string='Products')


    @api.constrains('date_start')
    def _code_constrains(self):
        if self.date_start > self.date_end:
            raise ValidationError(_("'Start Date' must be before 'End Date'"))
    # Get the product and category filters 

    def get_product(self):
        product_pool=self.env['product.product']
        # pdb.set_trace()
        if not self.category_id :
            product_ids = product_pool.search([('type','!=','service')])
            return product_ids
        # elif self.product_ids:
        #     product_ids = product_pool.search([('id','in',self.product_ids.ids)])

        #     return product_ids 
        elif self.category_id:
            product_ids = product_pool.search([('categ_id','in',self.category_id.ids),('type','!=','service')])

            return product_ids

    def generate_report(self):

    
        # #Create Workbook and Worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "GRN"
        report_heading = " GRN Register from" + ' ' + datetime.strftime(self.date_start, '%d-%m-%Y') + ' '+ 'To' + ' '+ datetime.strftime( self.date_end, '%d-%m-%Y')
        # ws.column_dimensions['B'].width = 20
        # ws.column_dimensions['c'].width = 25
        # ws.column_dimensions['D'].width = 25
        # ws.column_dimensions['E'].width = 20
        # ws.column_dimensions['F'].width = 25
        # ws.column_dimensions['G'].width = 25
        # ws.column_dimensions['I'].width = 25
        # ws.column_dimensions['J'].width = 15
        # ws.column_dimensions['K'].width = 15
        # ws.column_dimensions['L'].width = 15
        # # current_row=4
        # #Border
        thin = Side(border_style="thin", color="000000")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=18)

        comapny = ws.cell(row=1, column=1, value=self.env.user.company_id.name + report_heading)
        comapny.alignment = Alignment(horizontal='center',vertical='center')
        comapny.border = Border(bottom=thin,top=thin)
        comapny.font = Font(size=10,name='Calibri')


        # Defining the Table Caolumn headings
        sl_no = ws.cell(row=3, column=1, value="S.No")
        sl_no.alignment = copy(comapny.alignment)
        product_cat = ws.cell(row=3, column=2, value="Product Category (Brand)")
        product_cat.alignment = copy(comapny.alignment)
        grn_date = ws.cell(row=3, column=3, value="GR Date")
        grn_date.alignment = copy(comapny.alignment)
        # region = ws.cell(row=3, column=4, value="Region")
        # region.alignment = Alignment(horizontal='center',vertical='center')
        invoice_date = ws.cell(row=3, column=4, value="Invoice Date")
        invoice_date.alignment = Alignment(horizontal='center',vertical='center')
        grn_no = ws.cell(row=3, column=5, value="GRN No")
        grn_no.alignment = Alignment(horizontal='center',vertical='center')
        invoice_no = ws.cell(row=3, column=6, value="Invoice No")
        invoice_no.alignment = Alignment(horizontal='center',vertical='center')
        created_by = ws.cell(row=3, column=7, value="Create By")
        created_by.alignment = Alignment(horizontal='center',vertical='center')
        grn_state = ws.cell(row=3, column=8, value="GR State")
        grn_state.alignment = Alignment(horizontal='center',vertical='center')

        supplier_name = ws.cell(row=3, column=9, value="Supplier Name")
        supplier_name.alignment = Alignment(horizontal='center',vertical='center')
        
        warehouse_name = ws.cell(row=3, column=10, value="Site Name")
        warehouse_name.alignment = Alignment(horizontal='center',vertical='center')
        # supplier = ws.cell(row=3, column=12, value="Supplier")
        # supplier.alignment = Alignment(horizontal='center',vertical='center')
        po_name = ws.cell(row=3, column=11, value="PO No")
        po_name.alignment = Alignment(horizontal='center',vertical='center')

        po_date = ws.cell(row=3, column=12, value="PO Order Date")
        po_date.alignment = Alignment(horizontal='center',vertical='center')
        product_id = ws.cell(row=3, column=13, value="Item Name")
        product_id.alignment = Alignment(horizontal='center',vertical='center')
        product_code = ws.cell(row=3, column=14, value="Dish Code")
        product_code.alignment = Alignment(horizontal='center',vertical='center')
        po_qty = ws.cell(row=3, column=15, value="PO Qty")
        po_qty.alignment = Alignment(horizontal='center',vertical='center')

        po_uom = ws.cell(row=3, column=16, value="PO Unit")
        po_uom.alignment = Alignment(horizontal='center',vertical='center')

        po_price = ws.cell(row=3, column=17, value="PO Rate")
        po_price.alignment = Alignment(horizontal='center',vertical='center')

        grn_qty = ws.cell(row=3, column=18, value="GRN Qty")
        grn_qty.alignment = Alignment(horizontal='center',vertical='center')

        grn_uom = ws.cell(row=3, column=19, value="GRN Unit")
        grn_uom.alignment = Alignment(horizontal='center',vertical='center')

        grn_price = ws.cell(row=3, column=20, value="GR Rate")
        grn_price.alignment = Alignment(horizontal='center',vertical='center')

        grn_vale = ws.cell(row=3, column=21, value="GR Subtotal")
        grn_vale.alignment = Alignment(horizontal='center',vertical='center')

        gst = ws.cell(row=3, column=22, value="GST %")
        gst.alignment = Alignment(horizontal='center',vertical='center')

        # po_uom = ws.cell(row=3, column=23, value="PO Unit")
        # po_uom.alignment = Alignment(horizontal='center',vertical='center')

        gst_amt = ws.cell(row=3, column=23, value="GST ")
        gst_amt.alignment = Alignment(horizontal='center',vertical='center')

        cgst = ws.cell(row=3, column=24, value="CGST %")
        cgst.alignment = Alignment(horizontal='center',vertical='center')

        cgst_amt = ws.cell(row=3, column=25, value="CGST")
        cgst_amt.alignment = Alignment(horizontal='center',vertical='center')

        s_gst = ws.cell(row=3, column=26, value="SGST %")
        s_gst.alignment = Alignment(horizontal='center',vertical='center')

        s_gst_amt = ws.cell(row=3, column=27, value="SGST")
        s_gst_amt.alignment = Alignment(horizontal='center',vertical='center')

        igst = ws.cell(row=3, column=28, value="IGST %")
        igst.alignment = Alignment(horizontal='center',vertical='center')

        igst_amt = ws.cell(row=3, column=29, value="IGST")
        igst_amt.alignment = Alignment(horizontal='center',vertical='center')

        # gst = ws.cell(row=3, column=30, value="GST %")
        # gst.alignment = Alignment(horizontal='center',vertical='center')

        other_tax = ws.cell(row=3, column=30, value="Other Tax %")
        other_tax.alignment = Alignment(horizontal='center',vertical='center')

        other_tax_amt = ws.cell(row=3, column=31, value="Other Tax Amount")
        other_tax.alignment = Alignment(horizontal='center',vertical='center')

        grn_amt = ws.cell(row=3, column=32, value="GR Amount")
        grn_amt.alignment = Alignment(horizontal='center',vertical='center')

        grn_qty_uom = ws.cell(row=3, column=33, value="GRN qty by UOM")
        grn_qty_uom.alignment = Alignment(horizontal='center',vertical='center')

        default_uom = ws.cell(row=3, column=34, value="Default UOM")
        default_uom.alignment = Alignment(horizontal='center',vertical='center')



        
        # get the Extenal PO Info
        query='''
        select 
        sp.id as picking_id,
        sp.name as grn_name,
        sp.date_done,
        sp.state ,
        sw.name as w_name ,
        ru.display_name as u_name,
        sm.product_id as m_product,
        sm.product_id as po_product,
        sm.name,
        sm.id as sm_id,
        sm.product_uom_qty as grn_qty,
        sm.price_unit as grn_price,
        uu.name as grn_p_unit,
        pol.price_unit as po_price,
        sm.purchase_line_id as po_line_id,
        sp.bill_date as b_date,
        sp.bill_number as b_no

        from    uom_uom as uu,
                res_partner as ru,
                stock_picking_type as spt,
                stock_warehouse as sw,
                purchase_order_line as pol,
                stock_picking sp,
                stock_move sm,
                res_users as rsu
        
        where  sp.id = sm.picking_id and
        pol.id = sm.purchase_line_id and
        spt.id = sp.picking_type_id and 
        sw.id = spt.warehouse_id and
        rsu.partner_id =ru.id and
        rsu.id = sp.create_uid and
        uu.id = sm.product_uom and
        date_done > %s  and 
        date_done < %s and 
        sp.state = 'done' 
        '''

        params = (self.date_start ,self.date_end)

        self.env.cr.execute(query, params)
        result = self.env.cr.dictfetchall()
        

        query_sto = '''
        select 
        sp.id  as picking_id ,
        sp.name  as grn_name ,
        sp.date_done::date,sp.state ,
        sw.name as w_name ,
        ru.display_name as u_name,
        sm.product_id as m_product,
        sm.product_id as po_product,
        sm.name,
        sm.z_price as sm_zprice,
        sp.bill_date as b_date,
        sp.bill_number as b_no,
        sm.product_uom_qty as grn_qty,
        sm.price_unit as grn_price,
        sm.z_price as z_price,
        uu.name as grn_p_unit,
        sp.inter_company_transfer_id as sto_num,
        sm.id as sm_id,
        sp.id as sp_id,
        pt.name as it_name,
        um.name as u_nm,
        ictl.quantity as q_at,
        pt.default_code as pdf_cd,
        pc.name as pc_name,
        icte.processed_date::date as sto_date

    from  uom_uom as uu,
    res_partner as ru,
    stock_picking_type as spt,
    stock_warehouse as sw,  
    inter_company_transfer_ept as icte,
    stock_picking sp,
    stock_move sm,
    product_product as pp,
    product_template as pt,
    uom_uom as um,
    inter_company_transfer_line_ept as ictl,
    product_category as pc,
    res_users as rsu
        
        where  sp.id = sm.picking_id and
    icte.id= sp.inter_company_transfer_id and
        spt.id = sp.picking_type_id and 

        sw.id = spt.warehouse_id and
        rsu.id = sp.create_uid and
        uu.id = sm.product_uom and
        sp.date_done > %s  and 
        sp.date_done < %s and 
        sm.product_id = pp.id and
        pp.product_tmpl_id = pt.id and
        sp.picking_type_name = 'Receipts' and
        pt.uom_id = um.id and
        ictl.inter_company_transfer_id = sp.inter_company_transfer_id and
        ictl.product_id = sm.product_id and 
        pt.categ_id = pc.id and
        rsu.partner_id =ru.id and
        sp.state = 'done' 


        '''

        params = (self.date_start ,self.date_end)

        self.env.cr.execute(query_sto, params)
        result_sto = self.env.cr.dictfetchall()


        
        row_c=4
        sl_num=1
        if self.grn_type =='external':
            po_percentage = ws.cell(row=3, column=35, value="Po Fulfilment %")
            po_percentage.alignment = Alignment(horizontal='center',vertical='center')

            po_delivery_date = ws.cell(row=3, column=36, value="PO Delivery Date")
            po_delivery_date.alignment = Alignment(horizontal='center',vertical='center')            
            for grn_line in result:
                p_ids = [x.categ_id.id for x in self.get_product()]
                if grn_line['sm_id']:
                        sm_ids = self.env['stock.move.line'].search([('move_id','=',grn_line['sm_id']),('picking_id','=',grn_line['picking_id'])])
                # pdb.set_trace()
                pur_line_id = self.env['purchase.order.line'].search([('id','=',grn_line['po_line_id'])])
                if  pur_line_id.product_id.categ_id.id  in p_ids and sm_ids.picking_id.picking_type_id.code=='incoming':
                    sgst_rate= sgst_amount= cgst_rate= cgst_amount= total_tax_amount=other_rate=other_amount=igst_amount =igst_rate=sub_total_amount=gst_total=total_with_tax_amount= 0.0
                    if pur_line_id.taxes_id:
                        for each_line in pur_line_id.taxes_id:
                            if each_line.tax_group_id.name == 'IGST':
                                for each_tax in each_line:
                                    igst_rate  = each_tax.amount if each_tax.amount else 0.0
                            # elif each_line.tax_group_id.name != 'IGST' and each_line.tax_group_id.name != 'GST':
                            #     for each_tax in each_line:
                            #         other_rate  = each_tax.amount if each_tax.amount else ' '
                            elif each_line.tax_group_id.name == 'GST':
                                for each_tax in each_line.children_tax_ids:
                                    sgst_rate  = each_tax.amount if each_tax.amount else 0.0
                                    cgst_rate  = each_tax.amount if each_tax.amount else 0.0
                    sl_val = ws.cell(row=row_c, column=1, value=sl_num)
                    sl_val.alignment = Alignment(horizontal='center',vertical='center')
                    product_cat = ws.cell(row=row_c, column=2, value=pur_line_id.product_id.categ_id.complete_name)
                    grn_date = ws.cell(row=row_c, column=3, value=grn_line['date_done'] )
                    grn_date .alignment = Alignment(horizontal='center',vertical='center')

                    invoice_date = ws.cell(row=row_c, column=4, value=grn_line['b_date'])
                    invoice_date .alignment = Alignment(horizontal='center',vertical='center')

                    grn_no = ws.cell(row=row_c, column=5, value=grn_line['grn_name'])
                    grn_no.alignment = Alignment(horizontal='center',vertical='center')

                    invoice_no = ws.cell(row=row_c, column=6, value=grn_line['b_no'])
                    invoice_no.alignment = Alignment(horizontal='center',vertical='center')

                    
                    created_by = ws.cell(row=row_c, column=7, value=grn_line['u_name'])
                    created_by .alignment = Alignment(horizontal='center',vertical='center')

                    
                    grn_state = ws.cell(row=row_c, column=8, value=grn_line['state'])
                    grn_state.alignment = Alignment(horizontal='center',vertical='center')

                    supplier_name = ws.cell(row=row_c, column=9, value=pur_line_id.order_id.partner_id.name)
                    supplier_name .alignment = Alignment(horizontal='center',vertical='center')


                    warehouse_name = ws.cell(row=row_c, column=10, value= grn_line['w_name'])
                    po_name = ws.cell(row=row_c, column=11, value=pur_line_id.order_id.name)
                    po_date = ws.cell(row=row_c, column=12, value=pur_line_id.order_id.date_approve)
                    product_id = ws.cell(row=row_c, column=13, value=pur_line_id.product_id.name)
                    # # # print("current_prod['lot_name']current_prod['lot_name']current_prod['lot_name']",current_prod['lot_name'])
                    product_code = ws.cell(row=row_c, column=14, value=pur_line_id.product_id.default_code )
                    po_qty = ws.cell(row=row_c, column=15, value=pur_line_id.product_qty )
                    po_uom = ws.cell(row=row_c, column=16, value=pur_line_id.product_uom.name )
                    po_price = ws.cell(row=row_c, column=17, value=grn_line['po_price'])
                    grn_qty = ws.cell(row=row_c, column=18, value=grn_line['grn_qty'] )
                    
                    grn_uom = ws.cell(row=row_c, column=19, value=sm_ids.product_uom_id.name)
                    grn_price = ws.cell(row=row_c, column=20, value=grn_line['grn_price'])
                    if grn_line['grn_price']  and grn_line['grn_qty']:
                        g_v =grn_line['grn_price'] * grn_line['grn_qty']
                        cgst_amount =(g_v *cgst_rate )/100
                        sgst_amount =(g_v *cgst_rate )/100
                        igst_amount =(g_v *igst_rate )/100
                        other_amount =(g_v *other_rate )/100
                        total_tax_amount = cgst_amount+igst_amount+sgst_amount+other_amount
                    else:
                        g_v =0.0
                    grn_vale = ws.cell(row=row_c, column=21, value=(g_v))
                    gst = ws.cell(row=row_c, column=22, value=cgst_rate + sgst_rate)
                    gst = ws.cell(row=row_c, column=23, value=(sgst_amount+cgst_amount))
                    cgst = ws.cell(row=row_c, column=24, value=cgst_rate)
                    cgst_amt = ws.cell(row=row_c, column=25, value=(cgst_amount))
                    s_gst = ws.cell(row=row_c, column=26, value=sgst_rate)
                    s_gst_amt = ws.cell(row=row_c, column=27, value=(sgst_amount))
                    igst = ws.cell(row=row_c, column=28, value=igst_rate)
                    igst_amt = ws.cell(row=row_c, column=29, value=(igst_amount))


                    other_tax = ws.cell(row=row_c, column=30, value=other_rate)
                    other_tax_amt = ws.cell(row=row_c, column=31, value=(other_amount))
                    grn_amt = ws.cell(row=row_c, column=32, value=g_v+ total_tax_amount)
                    grn_qty_uom = ws.cell(row=row_c, column=33, value= sm_ids.move_id.quantity_done)
                    default_uom = ws.cell(row=row_c, column=34, value=pur_line_id.product_id.uom_po_id.name)
                    po_percentage = ws.cell(row=row_c, column=35, value= grn_line['grn_qty']*100/pur_line_id.product_qty )
                    po_delivery_date = ws.cell(row=row_c, column=36, value=pur_line_id.order_id.date_planned)

                    sl_num +=1
                    row_c +=1



        else:
            for sto in result_sto:
                p_ids = [x.categ_id.id for x in self.get_product()]
                picking_id = self.env['stock.picking'].search([('id','=',sto['sp_id'])])
                sto_id = self.env['inter.company.transfer.ept'].search([('id','=',sto['sto_num'])])
                # pdb.set_trace()
                if picking_id.picking_type_id.code == 'incoming':
                    # for sto_line in sto_id.inter_company_transfer_line_ids:
                    
                    if  picking_id.picking_type_id.code == 'incoming':
                        smov_ids = self.env['stock.move']
                        sm_ids  = self.env['stock.move.line'].search([('move_id','=',sto['sm_id']),('picking_id','=',sto['picking_id'])])
                        if len(sm_ids)>1:
                            tot_qty= 0.0
                            for each in sm_ids:
                                    tot_qty += each.qty_done
                                # tot_qty = tot
                            product_name = each.product_id.name
                            po_uom_1 = each.product_uom_id.name
                            grn_qty_1 = each.move_id.quantity_done
                            po_uom_default = each.product_id.uom_po_id.name

                        else:
                            tot_qty = sm_ids.qty_done
                            po_uom_1 = sm_ids.product_uom_id.name
                            grn_qty_1 = sm_ids.move_id.quantity_done
                            po_uom_default = sm_ids.product_id.uom_po_id.name


                                # pdb.set_trace()
                        g_v=sgst_rate= sgst_amount= cgst_rate= cgst_amount= total_tax_amount=other_rate=other_amount=igst_amount =igst_rate=sub_total_amount=gst_total=total_with_tax_amount= 0.0
                            # if pur_line_id.taxes_id:
                            #     for each_line in pur_line_id.taxes_id:
                            #         if each_line.tax_group_id.name == 'IGST':
                            #             for each_tax in each_line:
                            #                 igst_rate  = each_tax.amount if each_tax.amount else ' '
                            #         elif each_line.tax_group_id.name != 'IGST' and each_line.tax_group_id.name != 'GST':
                            #             for each_tax in each_line:
                            #                 other_rate  = each_tax.amount if each_tax.amount else ' '
                            #         elif each_line.tax_group_id.name == 'GST':
                            #             for each_tax in each_line.children_tax_ids:
                            #                 sgst_rate  = each_tax.amount if each_tax.amount else ' '
                            #                 cgst_rate  = each_tax.amount if each_tax.amount else ' '
                        sl_val = ws.cell(row=row_c, column=1, value=sl_num)
                        sl_val.alignment = Alignment(horizontal='center',vertical='center')
                        product_cat = ws.cell(row=row_c, column=2, value=sto['pc_name'])
                        grn_date = ws.cell(row=row_c, column=3, value=sto['date_done'] )
                        grn_date .alignment = Alignment(horizontal='center',vertical='center')

                        invoice_date = ws.cell(row=row_c, column=4, value=sto['b_date'])
                        invoice_date .alignment = Alignment(horizontal='center',vertical='center')

                        grn_no = ws.cell(row=row_c, column=5, value=sto['grn_name'])
                        grn_no.alignment = Alignment(horizontal='center',vertical='center')

                        invoice_no = ws.cell(row=row_c, column=6, value=sto['b_no'])
                        invoice_no.alignment = Alignment(horizontal='center',vertical='center')

                            
                        created_by = ws.cell(row=row_c, column=7, value=sto['u_name'])
                        created_by .alignment = Alignment(horizontal='center',vertical='center')

                            
                        grn_state = ws.cell(row=row_c, column=8, value=sto['state'])
                        grn_state.alignment = Alignment(horizontal='center',vertical='center')

                        supplier_name = ws.cell(row=row_c, column=9, value=picking_id.partner_id.name)
                        supplier_name .alignment = Alignment(horizontal='center',vertical='center')


                        warehouse_name = ws.cell(row=row_c, column=10, value= sto['w_name'])
                        po_name = ws.cell(row=row_c, column=11, value=sto_id.name)
                        po_date = ws.cell(row=row_c, column=12, value=sto['sto_date'])
                        product_id = ws.cell(row=row_c, column=13, value=sto['it_name'])
                            # # # # print("current_prod['lot_name']current_prod['lot_name']current_prod['lot_name']",current_prod['lot_name'])
                        product_code = ws.cell(row=row_c, column=14, value=sto['pdf_cd'] )
                        po_qty = ws.cell(row=row_c, column=15, value=sto['q_at'])
                        po_uom = ws.cell(row=row_c, column=16, value=sto['u_nm'] )
                        po_price = ws.cell(row=row_c, column=17, value=sto['sm_zprice'])
                        grn_qty = ws.cell(row=row_c, column=18, value=sto['grn_qty'] )
                            # if sto_line['sm_id']:
                                
                            #     sm_ids = self.env['stock.move.line'].search([('move_id','=',sto_line['sm_id'])])
                        grn_uom = ws.cell(row=row_c, column=19, value=po_uom_1)
                        grn_price = ws.cell(row=row_c, column=20, value=sto['sm_zprice'])
                            # # if grn_line['grn_price']  and grn_line['grn_qty']:
                            # #     g_v =grn_line['grn_price'] * grn_line['grn_qty']
                            # #     cgst_amount =(g_v *cgst_rate )/100
                            # #     sgst_amount =(g_v *cgst_rate )/100
                            # #     igst_amount =(g_v *igst_rate )/100
                            # #     other_amount =(g_v *other_rate )/100
                            # #     total_tax_amount = cgst_amount+igst_amount+sgst_amount+other_amount
                            # # else:
                            # #     g_v =0.0
                        grn_vale = ws.cell(row=row_c, column=21, value=sto['grn_qty']*sto['sm_zprice'])
                        gst = ws.cell(row=row_c, column=22, value=cgst_rate + sgst_rate)
                        gst = ws.cell(row=row_c, column=23, value=(sgst_amount+cgst_amount))
                        cgst = ws.cell(row=row_c, column=24, value=cgst_rate)
                        cgst_amt = ws.cell(row=row_c, column=25, value=(cgst_amount))
                        s_gst = ws.cell(row=row_c, column=26, value=sgst_rate)
                        s_gst_amt = ws.cell(row=row_c, column=27, value=(sgst_amount))
                        igst = ws.cell(row=row_c, column=28, value=igst_rate)
                        igst_amt = ws.cell(row=row_c, column=29, value=(igst_amount))


                            # other_tax = ws.cell(row=row_c, column=30, value=other_rate)
                            # other_tax_amt = ws.cell(row=row_c, column=31, value=(other_amount))
                        grn_amt = ws.cell(row=row_c, column=32, value=(sto['grn_qty']*sto['sm_zprice'])+(sgst_amount+cgst_amount)+cgst_amount)
                        grn_qty_uom = ws.cell(row=row_c, column=33, value=sm_ids.move_id.quantity_done)
                        default_uom = ws.cell(row=row_c, column=34, value=sto['u_nm'])
                            # po_percentage = ws.cell(row=row_c, column=35, value= tot_qty*100/sto_line.quantity )
                            # po_delivery_date = ws.cell(row=row_c, column=36, value=picking_id.date_done)

                        sl_num +=1
                        row_c +=1








        fp = io.BytesIO()
        wb.save(fp)
        excel_file = base64.encodestring(fp.getvalue())
        self.grn_report = excel_file
        self.grn_printed = True
        self.file_name = "GRN.xlsx"
        fp.close()

        return {
        'view_mode': 'form',
        'res_id': self.id,
        'res_model': 'grn.report',
        'view_type': 'form',
        'type': 'ir.actions.act_window',
        'context': self.env.context,
        'target': 'new',
                   }


   
