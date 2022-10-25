# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.



import io
import locale
import base64
import textwrap
from copy import copy

from datetime import datetime
from openpyxl import Workbook
from odoo import models, fields, api,_
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from odoo.exceptions import UserError, ValidationError,Warning
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
import pdb

class MrReport(models.TransientModel):
    _name = "mr.report"
    _description = "mr Report"

    date_start = fields.Datetime(string="Start Date", required=True)
    date_end = fields.Datetime(string="End Date", required=True, default=fields.Datetime.now)
    mr_report = fields.Binary('MR Report')
    file_name = fields.Char('File Name')
    # s_report_printed = fields.Boolean('MR Printed')
    mr_printed = fields.Boolean('Mr Printed')


    @api.constrains('date_start')
    def _code_constrains(self):
        if self.date_start > self.date_end:
            raise ValidationError(_("'Start Date' must be before 'End Date'"))


    def generate_report(self):
    
        #Create Workbook and Worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "MR"
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['c'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        # current_row=4
        #Border
        thin = Side(border_style="thin", color="000000")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=18)

        comapny = ws.cell(row=1, column=1, value=self.env.user.company_id.name)
        comapny.alignment = Alignment(horizontal='center',vertical='center')
        comapny.border = Border(bottom=thin,top=thin)
        comapny.font = Font(size=10,name='Calibri')


        # Defining the Table Caolumn headings
        sl_no = ws.cell(row=3, column=1, value="S.No")
        sl_no.alignment = copy(comapny.alignment)
        mr_ref = ws.cell(row=3, column=2, value="MR Ref")
        mr_ref.alignment = copy(comapny.alignment)
        mr_date = ws.cell(row=3, column=3, value="Mr Date")
        mr_date.alignment = copy(comapny.alignment)
        # region = ws.cell(row=3, column=4, value="Region")
        # region.alignment = Alignment(horizontal='center',vertical='center')
        operation_type = ws.cell(row=3, column=4, value="Operation Type")
        operation_type.alignment = Alignment(horizontal='center',vertical='center')
        source = ws.cell(row=3, column=5, value="Source Location")
        source.alignment = Alignment(horizontal='center',vertical='center')
        destination = ws.cell(row=3, column=6, value="Destination Location")
        destination.alignment = Alignment(horizontal='center',vertical='center')
        product_id = ws.cell(row=3, column=7, value="Product")
        product_id.alignment = Alignment(horizontal='center',vertical='center')

        product_cat_id = ws.cell(row=3, column=8, value="Product Category")
        product_cat_id.alignment = Alignment(horizontal='center',vertical='center')

        Product_code = ws.cell(row=3, column=9, value="Eat Code")
        Product_code.alignment = Alignment(horizontal='center',vertical='center')

        demand_qty = ws.cell(row=3, column=10, value="Demand Qty")
        demand_qty.alignment = Alignment(horizontal='center',vertical='center')

        done_qty = ws.cell(row=3, column=11, value="Done Qty")
        done_qty.alignment = Alignment(horizontal='center',vertical='center')
        
        product_uom = ws.cell(row=3, column=12, value="Unit of Measure")
        product_uom.alignment = Alignment(horizontal='center',vertical='center')
        # supplier = ws.cell(row=3, column=12, value="Supplier")
        # supplier.alignment = Alignment(horizontal='center',vertical='center')
        state = ws.cell(row=3, column=13, value="State")
        state.alignment = Alignment(horizontal='center',vertical='center')
        product_cost = ws.cell(row=3, column=14, value="MAP")
        product_cost.alignment = Alignment(horizontal='center',vertical='center')
        subtotal = ws.cell(row=3, column=15, value="Subtotal")
        subtotal.alignment = Alignment(horizontal='center',vertical='center')
        create_by = ws.cell(row=3, column=16, value="Create By")
        create_by.alignment = Alignment(horizontal='center',vertical='center')
        # margin = ws.cell(row=3, column=14, value="Margins")
        # margin.alignment = Alignment(horizontal='center',vertical='center')
        # gm = ws.cell(row=3, column=15, value="GM%")
        # gm.alignment = Alignment(horizontal='center',vertical='center')
        # amount = ws.cell(row=3, column=16, value="Amount")
        # amount.alignment = Alignment(horizontal='center',vertical='center')

        # Seaech the account.move in the range of the from date and to date

        # query = '''select * from stock_indent_order_line as sirl
        # JOIN stock_indent_order sir ON sir.id = sirl.mrp_indent_product_line_id 
        # where  sir.approve_date > %s and sir.approve_date < %s and sir.state = 'done'
        # '''

        query = '''select distinct 
        sirl.id,
        sirl.product_id,
        sirl.done_qty,
        sirl.product_uom_qty ,
        sir.location_dest_id,
        uu.name as p_unit,
        sir.name as ref,
        sir.indent_date as date,
        sir.state,
        sir.create_uid,
        sl.complete_name as s_name,
        spt.name as p_type_name ,
        sw.name as w_name
        from 
            stock_location as sl,
            uom_uom as uu,
            stock_picking_type as spt,
            stock_warehouse as sw,
            stock_indent_order_line as sirl
        JOIN stock_indent_order sir ON sir.id = sirl.mrp_indent_product_line_id 

        where  

        sir.indent_date > %s and 
        sir.indent_date < %s and 
        sir.state = 'done' and 
        sirl.product_uom = uu.id   and  
        sl.id= sir.location_id  and 
        spt.id = sir.picking_type_id  and
        sw.id = spt.warehouse_id


        '''
        params = (self.date_start ,self.date_end)

        self.env.cr.execute(query, params)
        result = self.env.cr.dictfetchall()


        move_query=''' 
        select  
            product_id,
            sum(sml.qty_done) as done

        from 

            stock_picking as sp,
            stock_move_line as sml


        where 
            sp.origin =%s and 
            sml.product_id =%s and
            sml.picking_id = sp.id
            group by product_id'''
        

        row_c=4
        sl_num=1
        for indent_line in result:

            move_params = (indent_line['ref'] ,indent_line['product_id'])

            self.env.cr.execute(move_query, move_params)
            move_result = self.env.cr.dictfetchall()
            print("move_result[0]['done']move_result[0]['done']move_result[0]['done']",move_result)
            if move_result:
                if move_result[0]['done'] == 0:
                    done=0.0
                else:
                    done =move_result[0]['done']
            else:
                done=0.0



            p_quer='''
            select 
                pt.name as p_name,
                uu.name as uu_name,
                pc.complete_name  as pc_name,
                pt.default_code as p_code
            from  
                product_product as pp,
                product_template as pt,
                uom_uom as uu,
                product_category as pc
            where 
                product_tmpl_id = pt.id and 
                pp.id =%s and
                uu.id = pt.uom_id and
                pc.id = pt.categ_id
            '''
            self.env.cr.execute(p_quer, [indent_line['product_id']])
            product = self.env.cr.dictfetchall()

            sl_val = ws.cell(row=row_c, column=1, value=sl_num)
            sl_val.alignment = Alignment(horizontal='center',vertical='center')
            mr_ref = ws.cell(row=row_c, column=2, value=indent_line['ref'])
            
            mr_date = ws.cell(row=row_c, column=3, value=indent_line['date'] )
            mr_date .alignment = Alignment(horizontal='center',vertical='center')
            # # region_val = ws.cell(row=row_c, column=4, value="Check")
            # # region_val .alignment = Alignment(horizontal='center',vertical='center')
            operation_type = ws.cell(row=row_c, column=4, value=indent_line['w_name'] +':'+indent_line['p_type_name'])
            operation_type.alignment = Alignment(horizontal='center',vertical='center')
            source = ws.cell(row=row_c, column=5, value=indent_line['s_name'])
            source.alignment = Alignment(horizontal='center',vertical='center')
            if indent_line['location_dest_id']:
                s_id = indent_line['location_dest_id']
                sl_id = self.env['stock.location'].search([('id','=',s_id)])
            destination = ws.cell(row=row_c, column=6, value=sl_id.complete_name)
            destination .alignment = Alignment(horizontal='center',vertical='center')
            if indent_line['product_id']:
                c_p =indent_line['product_id']
                p_id = self.env['product.product'].search([('id','=',c_p)])
            product_id = ws.cell(row=row_c, column=7, value=product[0]['p_name'])
            product_id.alignment = Alignment(horizontal='center',vertical='center')

            product_cat_id = ws.cell(row=row_c, column=8, value=product[0]['pc_name'])
            product_cat_id.alignment = Alignment(horizontal='center',vertical='center')

            Product_code = ws.cell(row=row_c, column=9, value=product[0]['p_code'])
            Product_code .alignment = Alignment(horizontal='center',vertical='center')


            demand_qty = ws.cell(row=row_c, column=10, value= indent_line['product_uom_qty'])
            done_qty = ws.cell(row=row_c, column=11, value= done)
            product_uom = ws.cell(row=row_c, column=12, value=indent_line['p_unit'])
            # # supplier_val = ws.cell(row=row_c, column=12, value="Supplier")
            state = ws.cell(row=row_c, column=13, value=indent_line['state'])
            product_cost = ws.cell(row=row_c, column=14, value=p_id.standard_price)
            # # # print("current_prod['lot_name']current_prod['lot_name']current_prod['lot_name']",current_prod['lot_name'])
            subtotal = ws.cell(row=row_c, column=15, value=(indent_line['product_uom_qty']*p_id.standard_price) )
            amount_val = ws.cell(row=row_c, column=16, value=self.env['res.users'].search([('id','=',indent_line['create_uid'])]).name)
            # amount_val = ws.cell(row=row_c, column=16, value=product.quantity *product.price_unit )
            #     warehouse_val = ws.cell(row=row_c, column=19, value="Check")

            sl_num +=1
            row_c +=1








        fp = io.BytesIO()
        wb.save(fp)
        excel_file = base64.encodestring(fp.getvalue())
        self.mr_report = excel_file
        self.mr_printed = True
        self.file_name = "MR.xlsx"
        fp.close()

        return {
        'view_mode': 'form',
        'res_id': self.id,
        'res_model': 'mr.report',
        'view_type': 'form',
        'type': 'ir.actions.act_window',
        'context': self.env.context,
        'target': 'new',
                   }


   
