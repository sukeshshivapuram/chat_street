# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.



import io
import locale
import base64
import textwrap
from copy import copy

from datetime import datetime
from openpyxl import Workbook
from odoo import models, fields, api,_
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from odoo.exceptions import UserError, ValidationError,Warning
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
import pdb

class GRMismatchReport(models.TransientModel):
    _name = "gr.mismatch.report"
    _description = "gr.mismatch.report"

    date_start = fields.Datetime(string="Start Date", required=True)
    date_end = fields.Datetime(string="End Date", required=True, default=fields.Datetime.now)
    gr_report = fields.Binary('GR Mismatch Report')
    file_name = fields.Char('File Name')
    gr_printed = fields.Boolean('GR Printed')
    
    

    @api.constrains('date_start')
    def _code_constrains(self):
        if self.date_start > self.date_end:
            raise ValidationError(_("'Start Date' must be before 'End Date'"))
    # Get the product and category filters 

    

    def generate_report(self):

    
        # #Create Workbook and Worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "To-Gr Mismatch"
        report_heading = " To-Gr Mismatch from" + ' ' + datetime.strftime(self.date_start, '%d-%m-%Y') + ' '+ 'To' + ' '+ datetime.strftime( self.date_end, '%d-%m-%Y')
        
        # #Border
        thin = Side(border_style="thin", color="000000")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=18)

        comapny = ws.cell(row=1, column=1, value=self.env.user.company_id.name + report_heading)
        comapny.alignment = Alignment(horizontal='center',vertical='center')
        comapny.border = Border(bottom=thin,top=thin)
        comapny.font = Font(size=10,name='Calibri')


        # Defining the Table Caolumn headings
        sl_no = ws.cell(row=3, column=1, value="S.No")
        sl_no.alignment = copy(comapny.alignment)

        # date = ws.cell(row=3, column=2, value="Date")
        # date.alignment = copy(comapny.alignment)

        sto_date = ws.cell(row=3, column=2, value="Date")
        sto_date.alignment = copy(comapny.alignment)

        to_num = ws.cell(row=3, column=3, value="STO Number")
        to_num.alignment = Alignment(horizontal='center',vertical='center')

        dc_reference = ws.cell(row=3, column=4, value="DC REFERENCE")
        dc_reference.alignment = Alignment(horizontal='center',vertical='center')

        to_state = ws.cell(row=3, column=5, value="STO Status")
        to_state.alignment = Alignment(horizontal='center',vertical='center')

        supply_site = ws.cell(row=3, column=6, value="Supply to Site")
        supply_site.alignment = Alignment(horizontal='center',vertical='center')

        supply_name = ws.cell(row=3, column=7, value="Supplier Name")
        supply_name.alignment = Alignment(horizontal='center',vertical='center')

        product_cat = ws.cell(row=3, column=8, value="Item Category Name")
        product_cat.alignment = Alignment(horizontal='center',vertical='center')

        item_name = ws.cell(row=3, column=9, value="Item Name")
        item_name.alignment = Alignment(horizontal='center',vertical='center')
        
        p_unit = ws.cell(row=3, column=10, value="Unit")
        p_unit.alignment = Alignment(horizontal='center',vertical='center')

        # supplier = ws.cell(row=3, column=12, value="Supplier")
        # supplier.alignment = Alignment(horizontal='center',vertical='center')

        to_qty = ws.cell(row=3, column=11, value="DC Quantity")
        to_qty.alignment = Alignment(horizontal='center',vertical='center')

        to_rate = ws.cell(row=3, column=12, value="Rate")
        to_rate.alignment = Alignment(horizontal='center',vertical='center')

        to_amt = ws.cell(row=3, column=13, value="TO Amt")
        to_amt.alignment = Alignment(horizontal='center',vertical='center')

        gr_qty = ws.cell(row=3, column=14, value="GR Qty")
        gr_qty.alignment = Alignment(horizontal='center',vertical='center')

        grn_rate = ws.cell(row=3, column=15, value="GR Rate")
        grn_rate.alignment = Alignment(horizontal='center',vertical='center')

        gr_amt = ws.cell(row=3, column=16, value="GR Amt")
        gr_amt.alignment = Alignment(horizontal='center',vertical='center')

        diff_qty = ws.cell(row=3, column=17, value="Diff Qty")
        diff_qty.alignment = Alignment(horizontal='center',vertical='center')

        diff_amt = ws.cell(row=3, column=18, value="Diff Amt")
        diff_amt.alignment = Alignment(horizontal='center',vertical='center')

        city = ws.cell(row=3, column=19, value="City")
        city.alignment = Alignment(horizontal='center',vertical='center')

        dc_return = ws.cell(row=3, column=20, value="DC return")       
        dc_return.alignment = Alignment(horizontal='center',vertical='center')
        
        return_qty = ws.cell(row=3, column=21, value="Return Quantity")
        return_qty.alignment = Alignment(horizontal='center',vertical='center')

 
        # query_ret = '''
        # select 
        #     smr.reference as smr_ref,
        #     smr.picking_id as smr_dcreturn,
        #     smr.product_uom_qty as smr_qty
        # from
        #     stock_move as smr,
        #     stock_picking as spr
        # where
        #     smr.picking_id = spr.id           

        # '''

        # self.env.cr.execute(query_ret)
        # ret_result = self.env.cr.dictfetchall()


        query_sto = '''
        select  
            ict.id, 
            ict.name as ict_name,
            ict.processed_date as ict_date,
            sp.id,
            sp.name 
        from    
            stock_picking sp,
            inter_company_transfer_ept as ict 
        where
        
            sp.inter_company_transfer_id =ict.id  and
            ict.processed_date > %s  and 
            ict.processed_date < %s and
            ict.state = 'processed' 

        '''

        params = (self.date_start ,self.date_end)

        self.env.cr.execute(query_sto, params)
        result = self.env.cr.dictfetchall()


        
        row_c=4
        sl_num=1

        # for grn_line in result:
        EPT = self.env['inter.company.transfer.ept']
        ict_ids =EPT.search([('processed_date', '>',  self.date_start) , ('processed_date', '<', self.date_end)])
        if ict_ids:
            for each_ict in ict_ids:

                # Ge the GRN product quantity group by product
                # sp.picking_type_name = 'Receipts
                query ='''
                        select 
                            sm.product_id as p_id,
                            sm.origin_returned_move_id as ret_id,
                            sm.reference as sm_ref,
                            sm.product_qty as grn_qty,
                            sm.id as sm_id,
                            sp.id as sp_id
                        from    
                            stock_picking as sp,
                            stock_move as sm,
                            stock_picking_type as spt
                        where
                            sp.inter_company_transfer_id =%s and 
                            sm.picking_id = sp.id and
                            sp.picking_type_id = spt.id and
                            sp.state='done' and
                            spt.code = 'outgoing' and
                            sm.state='done'
                           
                            
                
                        '''


                # params = (each_ict.ids)

                self.env.cr.execute(query, each_ict.ids)
                rel_result = self.env.cr.dictfetchall()
                print(result,"seifuwiwe")






                if rel_result:
                    for each_sm in rel_result:
                        # if each_sm['ret_id']:

                        ret_quer='''
                        select 
                            sm.reference,sm.origin_returned_move_id ,sml.qty_done
                        from 
                            stock_move as sm,
                            stock_move_line as sml

                        where 
                            %s = origin_returned_move_id and
                            sm.id = sml.move_id
                  
                                '''


                        self.env.cr.execute(ret_quer, [each_sm['sm_id']])
                        dc_ret_result = self.env.cr.dictfetchall()
                        # print(dc_ret_result,"srgiuehi")


                        grqty_quer='''
                        select
                            
                            sml.qty_done as gr_qty
                        
                        from 
                            stock_move_line as sml,
                            
                            stock_picking as sp
                        
                        where
                            %s = sp.ict_id_no and
                            sml.picking_id= sp.id and
                            sml.product_id=%s


                            '''
                        self.env.cr.execute(grqty_quer, [each_sm['sp_id'],each_sm['p_id']]) 
                        gr_qty_result = self.env.cr.dictfetchall()
                        # print(gr_qty_result,"reteeetre")           
                        # Get the product name,uom and the product Categery
                        p_quer='''
                        select 
                                pt.name as p_name,
                                uu.name as uu_name,
                                pc.complete_name  as pc_name
                        from  
                            product_product as pp,
                            product_template as pt,
                            uom_uom as uu,
                            product_category as pc
                        where 
                            product_tmpl_id = pt.id and 
                            pp.id =%s and
                            uu.id = pt.uom_id and
                            pc.id = pt.categ_id
                            '''
                        self.env.cr.execute(p_quer, [each_sm['p_id']])
                        product = self.env.cr.dictfetchall()
                        print("product,aefwgei")


                        # Get the GRN Quantity from the STO
                        sto_query='''
                        select  
                                quantity as s_qty
                        from 
                                inter_company_transfer_line_ept 
                        where 
                                inter_company_transfer_id = %s and
                                product_id=%s
                                '''
                        params_s=[each_ict.id,each_sm['p_id']]
                        self.env.cr.execute(sto_query,params_s )
                        sto_result = self.env.cr.dictfetchall()
                        print(sto_result,"eifuwi")

                        curr_prod = self.env['product.product'].search([('id','=',each_sm['p_id'])])
                        # if curr_prod:
                        #     tax_per= 0.0
                        #     for each_tax in curr_prod.supplier_taxes_id:
                        #         # pdb.set_trace()
                        #         if each_tax.tax_group_id.name == "GST" :
                        #             for each_gst in each_tax.children_tax_ids:
                        #                 tax_per+= each_gst.amount
                        #         elif  each_tax.tax_group_id.name == "IGST":
                        #             tax_per+= each_tax.amount


                        sl_val = ws.cell(row=row_c, column=1, value=sl_num)
                        sl_val.alignment = Alignment(horizontal='center',vertical='center')


                        grn_date = ws.cell(row=row_c, column=2, value=each_ict.processed_date )
                        grn_date .alignment = Alignment(horizontal='center',vertical='center')

                        to_num = ws.cell(row=row_c, column=3, value=each_ict.name)
                        to_num .alignment = Alignment(horizontal='center',vertical='center')

                        dc_reference = ws.cell(row=row_c, column=4, value=each_sm['sm_ref'])
                       

                        to_state = ws.cell(row=row_c, column=5, value=each_ict.state)
                     
                        supply_site = ws.cell(row=row_c, column=6, value=each_ict.destination_warehouse_id.name)
                        supply_site.alignment = Alignment(horizontal='center',vertical='center')


                        supply_name = ws.cell(row=row_c, column=7, value=each_ict.source_warehouse_id.name)
                        supply_name .alignment = Alignment(horizontal='center',vertical='center')


                        product_cat = ws.cell(row=row_c, column=8, value=product[0]['pc_name'])


                        item_name = ws.cell(row=row_c, column=9, value=product[0]['p_name'])



                        p_unit = ws.cell(row=row_c, column=10, value= product[0]['uu_name'])

                        to_qty = ws.cell(row=row_c, column=11, value=each_sm['grn_qty'] )

                        to_rate = ws.cell(row=row_c, column=12, value=curr_prod.standard_price)

                        to_amt = ws.cell(row=row_c, column=13, value=each_sm['grn_qty']*curr_prod.standard_price)


                        gr_qty = ws.cell(row=row_c, column=14, value=gr_qty_result[0]['gr_qty'] if gr_qty_result else 0)

                        grn_rate = ws.cell(row=row_c, column=15, value= curr_prod.standard_price)

                        gr_amt = ws.cell(row=row_c, column=16, value=curr_prod.standard_price* gr_qty_result[0]['gr_qty'] if gr_qty_result else 0)

                        diff_qty = ws.cell(row=row_c, column=17, value=each_sm['grn_qty']-gr_qty_result[0]['gr_qty'] if gr_qty_result else each_sm['grn_qty'])

                        diff_amt = ws.cell(row=row_c, column=18, value=(each_sm['grn_qty']-gr_qty_result[0]['gr_qty'])*curr_prod.standard_price if gr_qty_result else 0)

                        city = ws.cell(row=row_c, column=19, value= "True" if each_ict.destination_warehouse_id.city_id.id == each_ict.source_warehouse_id.city_id.id else "False" )
        
                        dc_return = ws.cell(row=row_c, column=20, value=dc_ret_result[0]['reference'] if dc_ret_result else '' )       

                        return_qty = ws.cell(row=row_c, column=21, value=dc_ret_result[0]['qty_done'] if dc_ret_result else 0 )                            

                        sl_num +=1
                        row_c +=1






        


        fp = io.BytesIO()
        wb.save(fp)
        excel_file = base64.encodestring(fp.getvalue())
        self.gr_report = excel_file
        self.gr_printed = True
        self.file_name = "To-Gr Mismatch.xlsx"
        fp.close()

        return {
        'view_mode': 'form',
        'res_id': self.id,
        'res_model': 'gr.mismatch.report',
        'view_type': 'form',
        'type': 'ir.actions.act_window',
        'context': self.env.context,
        'target': 'new',
                   }


   
