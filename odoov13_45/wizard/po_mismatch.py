# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.



import io
import locale
import base64
import textwrap
from copy import copy

from datetime import datetime
from openpyxl import Workbook
from odoo import models, fields, api,_
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from odoo.exceptions import UserError, ValidationError,Warning
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
import pdb

class PoMismatch(models.TransientModel):
    _name = "po.mismatch"
    _description = "po.mismatch"

    date_start = fields.Datetime(string="Start Date", required=True)
    date_end = fields.Datetime(string="End Date", required=True, default=fields.Datetime.now)
    po_report = fields.Binary('PO Mismatch Report')
    file_name = fields.Char('File Name')
    po_printed = fields.Boolean('PO Mismatch Printed')
    
    

    @api.constrains('date_start')
    def _code_constrains(self):
        if self.date_start > self.date_end:
            raise ValidationError(_("'Start Date' must be before 'End Date'"))


    

    def generate_report(self):

    
        # #Create Workbook and Worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "PO-To Mismatch Report"
        report_heading = " PO-To Mismatch Report from" + ' ' + datetime.strftime(self.date_start, '%d-%m-%Y') + ' '+ 'To' + ' '+ datetime.strftime( self.date_end, '%d-%m-%Y')
        
        # #Border
        thin = Side(border_style="thin", color="000000")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=18)

        comapny = ws.cell(row=1, column=1, value=self.env.user.company_id.name + report_heading)
        comapny.alignment = Alignment(horizontal='center',vertical='center')
        comapny.border = Border(bottom=thin,top=thin)
        comapny.font = Font(size=10,name='Calibri')


        # Defining the Table Caolumn headings
        sl_no = ws.cell(row=3, column=1, value="S.No")
        sl_no.alignment = copy(comapny.alignment)


        to_date = ws.cell(row=3, column=2, value="To Date")
        to_date.alignment = copy(comapny.alignment)

        item_name = ws.cell(row=3, column=3, value="Item Name")
        item_name.alignment = Alignment(horizontal='center',vertical='center')

        p_unit = ws.cell(row=3, column=4, value="Measurement Unit")
        p_unit.alignment = Alignment(horizontal='center',vertical='center')

        po_site_name = ws.cell(row=3, column=5, value="PO Site Name")
        po_site_name.alignment = Alignment(horizontal='center',vertical='center')

        to_site_name = ws.cell(row=3, column=6, value="To Site Name")
        to_site_name.alignment = Alignment(horizontal='center',vertical='center')

        po_qty = ws.cell(row=3, column=7, value="PO Quantity")
        po_qty.alignment = Alignment(horizontal='center',vertical='center')

        to_qty = ws.cell(row=3, column=8, value="TO Quantity")
        to_qty.alignment = Alignment(horizontal='center',vertical='center')


        sto_num = ws.cell(row=3, column=9, value="PO NO")
        sto_num.alignment = Alignment(horizontal='center',vertical='center')

        
        to_no = ws.cell(row=3, column=10, value="TO NO")
        to_no.alignment = Alignment(horizontal='center',vertical='center')
       
        
        
        

       

        
        row_c=4
        sl_num=1

        # for grn_line in result:
        EPT = self.env['inter.company.transfer.ept']
        StockPicking = self.env['stock.picking']

        sto_rec= '''
                select  
                sp.id as dc_id,
                sp.name as dc_name,
                sp.state as dc_state,
                ict.id as ict_id,
                ict.name as sto_name,
                ict.source_warehouse_id as sw_id,
                ict.destination_warehouse_id as dw_id,
                ict.processed_date as ict_date
        from  
                stock_picking as sp,
                inter_company_transfer_ept ict 

        where   
            ict.processed_date > %s and 
            ict.processed_date < %s and 
            sp.inter_company_transfer_id  = ict.id and 
            sp.picking_type_name = 'Delivery Orders'
                '''
        params = (self.date_start,self.date_end)
        self.env.cr.execute(sto_rec, params)
        sto_result = self.env.cr.dictfetchall()

        

        for each_picking in sto_result:

            if each_picking:
                # Ge the GRN product quantity group by product
                query ='''
                select  *

                from    
                stock_move_line 
                where 
                picking_id = %s and
                state = 'done'
                '''
                params = (each_picking['dc_id'],)
                # 
                self.env.cr.execute(query, params)
                result = self.env.cr.dictfetchall()


                for each_product in result:
                    p_quer='''
                    select 
                        pt.name as p_name,
                        uu.name as uu_name,
                        pc.complete_name  as pc_name
                    from  
                        product_product as pp,
                        product_template as pt,
                        uom_uom as uu,
                        product_category as pc
                    where 
                        product_tmpl_id = pt.id and 
                        pp.id =%s and
                        uu.id = pt.uom_id and
                        pc.id = pt.categ_id
                    '''
                    self.env.cr.execute(p_quer, [each_product['product_id']])
                    product = self.env.cr.dictfetchall()




                    # Get the warehouse names 
                    sw_query='''
                    select  
                        *
                    from 
                        stock_warehouse 
                    where 
                        id = %s 
                        '''
                    params_sw=[each_picking['sw_id']]
                    params_dw=[each_picking['dw_id']]

                    self.env.cr.execute(sw_query,params_sw )
                    sw_result = self.env.cr.dictfetchall()

                    self.env.cr.execute(sw_query,params_dw )
                    dw_result = self.env.cr.dictfetchall()
                    # get the STO demand qty
                    ict_line_query ='''
                    select  *

                    from    
                    inter_company_transfer_line_ept 
                    where 
                    inter_company_transfer_id = %s and
                    product_id = %s
                    '''
                    params = (each_picking['ict_id'],each_product['product_id'])
                    # 
                    self.env.cr.execute(ict_line_query, params)
                    result = self.env.cr.dictfetchall()

                    print("resultresultresultresultresult",result)
                    sl_val = ws.cell(row=row_c, column=1, value=sl_num)
                    sl_val.alignment = Alignment(horizontal='center',vertical='center')


                    to_date = ws.cell(row=row_c, column=2, value= each_picking['ict_date'])
                    # to_date .alignment = Alignment(horizontal='center',vertical='center')

                    item_name = ws.cell(row=row_c, column=3, value=product[0]['p_name'])
                    # item_name .alignment = Alignment(horizontal='center',vertical='center')

                    p_unit = ws.cell(row=row_c, column=4, value=product[0]['uu_name'])
                    # dc_state.alignment = Alignment(horizontal='center',vertical='center')

                    po_site_name = ws.cell(row=row_c, column=5, value=dw_result[0]['name'])
                    # # sto_name.alignment = Alignment(horizontal='center',vertical='center')

                    to_site_name = ws.cell(row=row_c, column=6, value=sw_result[0]['name'])
                    # # supply_site.alignment = Alignment(horizontal='center',vertical='center')


                    po_qty = ws.cell(row=row_c, column=7, value=result[0]['quantity'])

                    to_qty = ws.cell(row=row_c, column=8, value=each_product['qty_done'])


                    sto_num = ws.cell(row=row_c, column=9, value=each_picking['sto_name'])



                    to_no = ws.cell(row=row_c, column=10, value= each_picking['dc_name'])

                    


                    sl_num +=1
                    row_c +=1



        


        fp = io.BytesIO()
        wb.save(fp)
        excel_file = base64.encodestring(fp.getvalue())
        self.po_report = excel_file
        self.po_printed = True
        self.file_name = "PO-TO Mismatch.xlsx"
        fp.close()

        return {
        'view_mode': 'form',
        'res_id': self.id,
        'res_model': 'po.mismatch',
        'view_type': 'form',
        'type': 'ir.actions.act_window',
        'context': self.env.context,
        'target': 'new',
                   }


   
