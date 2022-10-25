# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.



import io
import locale
import base64
import textwrap
from copy import copy

from datetime import datetime
from openpyxl import Workbook
from odoo import models, fields, api,_
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from odoo.exceptions import UserError, ValidationError,Warning
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
import pdb

class OpencloseReport(models.TransientModel):
    _name = "openclose.report"
    _description = "openclose.report"

    date_start = fields.Datetime(string="Start Date", required=True)
    date_end = fields.Datetime(string="End Date", required=True, default=fields.Datetime.now)
    openclose_report = fields.Binary('Stock Opening Closing Report')
    file_name = fields.Char('File Name')
    openclose_printed = fields.Boolean('Stock Opening Closing Printed')
    warehouse_id = fields.Many2many('stock.warehouse', )
    # pan_india = fields.Boolean('Pan India')


    @api.constrains('date_start')
    def _code_constrains(self):
        if self.date_start > self.date_end:
            raise ValidationError(_("'Start Date' must be before 'End Date'"))
    # Get the product and category filters 


    def generate_report(self):

    
        # #Create Workbook and Worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Opening and Closing"
        report_heading = " Stock Opening/Closing Report from" + ' ' + datetime.strftime(self.date_start, '%d-%m-%Y') + ' '+ 'To' + ' '+ datetime.strftime( self.date_end, '%d-%m-%Y')
        
        # #Border
        thin = Side(border_style="thin", color="000000")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=18)

        comapny = ws.cell(row=1, column=1, value=self.env.user.company_id.name + report_heading)
        comapny.alignment = Alignment(horizontal='center',vertical='center')
        comapny.border = Border(bottom=thin,top=thin)
        comapny.font = Font(size=10,name='Calibri')

        supply_site = ws.cell(row=3, column=1, value="Site Name")
        supply_site.alignment = copy(comapny.alignment)

        item_name = ws.cell(row=3, column=2, value="Item Name")
        item_name.alignment = Alignment(horizontal='center',vertical='center')

        p_unit = ws.cell(row=3, column=3, value="Unit of Measure")
        p_unit.alignment = Alignment(horizontal='center',vertical='center')

        sys_opening = ws.cell(row=3, column=4, value="Sys Opening Qty")
        sys_opening.alignment = Alignment(horizontal='center',vertical='center')

        sys_closing_qty = ws.cell(row=3, column=5, value="Sys Closing Qty")
        sys_closing_qty.alignment = Alignment(horizontal='center',vertical='center')

        dish_code= ws.cell(row=3, column=6, value="Internal Reference")
	    

        # caliculating the opening qty.
        opening_qty_query = '''
            select 
                    product_id ,
                    ABS(sum(product_qty)) as qty,
                    ABS(sum(value)) as value
            from 
                    inventory_base_report 
            where 
                    date < %s and 
                    product_id =%s and
                    warehouse_id = %s  and
                    transaction_types != 'internal'

            group by product_id

            '''

        # caliculating the closing qty.
        closing_qty_query = '''
            select 
                    product_id ,
                    ABS(sum(product_qty)) as qty,
                    ABS(sum(value)) as value
                    
                   
            from 
                    inventory_base_report

                 
            where 
                    date <= %s and 
                    product_id =%s and
                    warehouse_id = %s  and
                    transaction_types != 'internal' 
                   

            group by product_id

            '''


        row_c=4
        sl_num=1
	        
        for each_base in self.env['product.product'].search([]):

            # Get the product name,uom and the product Categery
            product_quer='''
            select 
                    pt.name as p_name,
                    uu.name as uu_name,
                    pc.complete_name  as pc_name
            from  
                product_product as pp,
                product_template as pt,
                uom_uom as uu,
                product_category as pc
            where 
                product_tmpl_id = pt.id and 
                pp.id =%s and
                uu.id = pt.uom_id and
                pc.id = pt.categ_id
                '''        
            self.env.cr.execute(product_quer, [each_base.id,])
            product = self.env.cr.dictfetchall() 


            # opening qty
            for trst in self.warehouse_id:

            	opening_qt_params = (self.date_start,each_base.id,trst.id)
            	self.env.cr.execute(opening_qty_query,opening_qt_params)
            	opening_qt_result = self.env.cr.dictfetchall()


            # Closing qty
            # for yuft in self.warehouse_id:

            	closing_qty_params = (self.date_end,each_base.id,trst.id)
            	self.env.cr.execute(closing_qty_query,closing_qty_params)
            	closing_qty_result = self.env.cr.dictfetchall()



            	p_op =opening_qt_result[0]['qty'] if opening_qt_result  else 0.0

            	p_clo=closing_qty_result[0]['qty'] if closing_qty_result  else 0.0

            	# w_name = closing_qty_result[0]['w_id'] if closing_qty_result[0]['w_id'] else None

            	supply_site = ws.cell(row=row_c, column=1, value=trst.name)

            	item_name = ws.cell(row=row_c, column=2, value=product[0]['p_name'])

            	p_unit = ws.cell(row=row_c, column=3, value=product[0]['uu_name'])

            	sys_opening = ws.cell(row=row_c, column=4, value=p_op)

            	sys_closing_qty = ws.cell(row=row_c, column=5, value=p_clo)

            	dish_code = ws.cell(row=row_c, column=6, value=each_base.default_code)

            	sl_num +=1
            	row_c +=1


        fp = io.BytesIO()
        wb.save(fp)
        excel_file = base64.encodestring(fp.getvalue())
        self.openclose_report = excel_file
        self.openclose_printed = True
        self.file_name = "Opening Closing Report.xlsx"
        fp.close()

        return {
        'view_mode': 'form',
        'res_id': self.id,
        'res_model': 'openclose.report',
        'view_type': 'form',
        'type': 'ir.actions.act_window',
        'context': self.env.context,
        'target': 'new',
                   }


