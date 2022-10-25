# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.



import io
import locale
import base64
import textwrap
from copy import copy

from datetime import datetime
from openpyxl import Workbook
from odoo import models, fields, api,_
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from odoo.exceptions import UserError, ValidationError,Warning
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
import pdb

class PurchaseReport(models.TransientModel):
    _name = "purchase.report.excel"
    _description = "purchase Report"

    date_start = fields.Datetime(string="Start Date", required=True)
    date_end = fields.Datetime(string="End Date", required=True, default=fields.Datetime.now)
    po_report = fields.Binary('PO Report')
    file_name = fields.Char('File Name')
    # s_report_printed = fields.Boolean('MR Printed')
    po_printed = fields.Boolean('PO Printed')
    # company_id = fields.Many2one('res.company',string='Company')
    # warehouse_ids = fields.Many2many('stock.warehouse',string='Warehouse',required="1")
    # location_id = fields.Many2one('stock.location',string='Location', domain="[('usage','!=','view')]")
    # production_id = fields.Many2one('stock.move',string='Product Id')
    start_date = fields.Date('Start Date')
    end_date = fields.Date('End Date')
    # grn_type = fields.Selection([('internal','Internal'),('external','External')],string='GRN Type')
    # category_id = fields.Many2many('product.category',string='Category')
    # product_ids = fields.Many2many('product.product',string='Products')


    @api.constrains('date_start')
    def _code_constrains(self):
        if self.date_start > self.date_end:
            raise ValidationError(_("'Start Date' must be before 'End Date'"))
    # Get the product and category filters 

    def get_product(self):
        product_pool=self.env['product.product']
        # pdb.set_trace()
        if not self.category_id :
            product_ids = product_pool.search([('type','!=','service')])
            return product_ids
        # elif self.product_ids:
        #     product_ids = product_pool.search([('id','in',self.product_ids.ids)])

        #     return product_ids 
        elif self.category_id:
            product_ids = product_pool.search([('categ_id','in',self.category_id.ids),('type','!=','service')])

            return product_ids

    def generate_report(self):

    
        # #Create Workbook and Worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Purchase Report"
        report_heading = " Purchase Register from" + ' ' + datetime.strftime(self.date_start, '%d-%m-%Y') + ' '+ 'To' + ' '+ datetime.strftime( self.date_end, '%d-%m-%Y')
        # ws.column_dimensions['B'].width = 20
        # ws.column_dimensions['c'].width = 25
        # ws.column_dimensions['D'].width = 25
        # ws.column_dimensions['E'].width = 20
        # ws.column_dimensions['F'].width = 25
        # ws.column_dimensions['G'].width = 25
        # ws.column_dimensions['I'].width = 25
        # ws.column_dimensions['J'].width = 15
        # ws.column_dimensions['K'].width = 15
        # ws.column_dimensions['L'].width = 15
        # # current_row=4
        # #Border
        thin = Side(border_style="thin", color="000000")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=18)

        comapny = ws.cell(row=1, column=1, value=self.env.user.company_id.name + report_heading)
        comapny.alignment = Alignment(horizontal='center',vertical='center')
        comapny.border = Border(bottom=thin,top=thin)
        comapny.font = Font(size=10,name='Calibri')


        # Defining the Table Caolumn headings
        sl_no = ws.cell(row=3, column=1, value="S.No")
        sl_no.alignment = copy(comapny.alignment)

        item_category = ws.cell(row=3, column=2, value="Item Category Name")
        item_category.alignment = copy(comapny.alignment)

        po_create_date = ws.cell(row=3, column=3, value="PO Creation Date")
        po_create_date.alignment = Alignment(horizontal='center',vertical='center')

        po_m_date = ws.cell(row=3, column=4, value="PO Modified DTTM")
        po_m_date.alignment = Alignment(horizontal='center',vertical='center')

        po_date = ws.cell(row=3, column=5, value="PO Date")
        po_date.alignment = Alignment(horizontal='center',vertical='center')

        po_deli_date = ws.cell(row=3, column=6, value="PO Delivery Date")
        po_deli_date.alignment = Alignment(horizontal='center',vertical='center')

        created_by = ws.cell(row=3, column=7, value="Created By")
        created_by.alignment = Alignment(horizontal='center',vertical='center')

        supplier_name = ws.cell(row=3, column=8, value="Supplier Name")
        supplier_name.alignment = Alignment(horizontal='center',vertical='center')
        
        warehouse_name = ws.cell(row=3, column=9, value="Site Name")
        warehouse_name.alignment = Alignment(horizontal='center',vertical='center')
        
        po_state = ws.cell(row=3, column=10, value="PO Status")
        po_state.alignment = Alignment(horizontal='center',vertical='center')

        po_name = ws.cell(row=3, column=11, value="PO Number")
        po_name.alignment = Alignment(horizontal='center',vertical='center')

        product_code = ws.cell(row=3, column=12, value="Dish Code")
        product_code.alignment = Alignment(horizontal='center',vertical='center')

        product_id = ws.cell(row=3, column=13, value="Item Name")
        product_id.alignment = Alignment(horizontal='center',vertical='center')

        po_qty = ws.cell(row=3, column=14, value="PO Qty")
        po_qty.alignment = Alignment(horizontal='center',vertical='center')

        po_uom = ws.cell(row=3, column=15, value="PO Unit")
        po_uom.alignment = Alignment(horizontal='center',vertical='center')

        po_price = ws.cell(row=3, column=16, value="PO Rate")
        po_price.alignment = Alignment(horizontal='center',vertical='center')


        po_value = ws.cell(row=3, column=17, value="Subtotal")
        po_value.alignment = Alignment(horizontal='center',vertical='center')

        cgst = ws.cell(row=3, column=18, value="CGST %")
        cgst.alignment = Alignment(horizontal='center',vertical='center')

        cgst_amt = ws.cell(row=3, column=19, value="CGST")
        cgst_amt.alignment = Alignment(horizontal='center',vertical='center')

        s_gst = ws.cell(row=3, column=20, value="SGST %")
        s_gst.alignment = Alignment(horizontal='center',vertical='center')

        s_gst_amt = ws.cell(row=3, column=21, value="SGST")
        s_gst_amt.alignment = Alignment(horizontal='center',vertical='center')

        igst = ws.cell(row=3, column=22, value="IGST %")
        igst.alignment = Alignment(horizontal='center',vertical='center')

        igst_amt = ws.cell(row=3, column=23, value="IGST")
        igst_amt.alignment = Alignment(horizontal='center',vertical='center')


        other_tax = ws.cell(row=3, column=24, value="Other Tax %")
        other_tax.alignment = Alignment(horizontal='center',vertical='center')

        other_tax_amt = ws.cell(row=3, column=25, value="Other Tax Amount")
        other_tax.alignment = Alignment(horizontal='center',vertical='center')

        po_tot = ws.cell(row=3, column=26, value="PO Amount")
        po_tot.alignment = Alignment(horizontal='center',vertical='center')



        qty_uom = ws.cell(row=3, column=27, value="Qty In UOM")
        qty_uom.alignment = Alignment(horizontal='center',vertical='center')

        default_uom = ws.cell(row=3, column=28, value="Default UOM")
        default_uom.alignment = Alignment(horizontal='center',vertical='center')

        rate_uom = ws.cell(row=3, column=29, value="Rate/UOM")
        rate_uom.alignment = Alignment(horizontal='center',vertical='center')
        
        # get the Extenal PO Info
        query='''
         select po.name  as po_name ,po.date_approve,po.state ,po.create_date,po.create_uid ,po.write_date ,pol.product_id as po_product,ru.display_name as p_name,
sw.name as w_name ,pol.id as po_line_id,po.date_planned as de_date


    from stock_picking_type as spt,stock_warehouse as sw,res_partner as ru, purchase_order po,purchase_order_line as pol
        
        where  pol.order_id = po.id and
        po.partner_id = ru.id and
        po.picking_type_id =spt.id and
        sw.id = spt.warehouse_id and
    
        po.date_approve > %s and 
        po.date_approve < %s and 
        po.state IN( 'done' ,'closed' ,'purchase' )
        '''

        params = (self.date_start ,self.date_end)

        self.env.cr.execute(query, params)
        result = self.env.cr.dictfetchall()
        

        # print("generate_reportgenerate_reportgenerate_reportgenerate_report",result)

        
        row_c=4
        sl_num=1

        for po_line in result:
            # p_ids = [x.categ_id.id for x in self.get_product()]
            # if grn_line['sm_id']:
            #         sm_ids = self.env['stock.move.line'].search([('move_id','=',grn_line['sm_id'])])
            pur_line_id = self.env['purchase.order.line'].search([('id','=',po_line['po_line_id'])])

            move_ids = self.env['stock.move'].search([('purchase_line_id','=',pur_line_id.id),('product_id','=',pur_line_id.product_id.id)])
            for each in move_ids:
                if each.picking_id.origin == pur_line_id.order_id.name:
                    move_id = each

            sgst_rate= sgst_amount= cgst_rate= cgst_amount= total_tax_amount=other_rate=other_amount=igst_amount =igst_rate=sub_total_amount=gst_total=total_with_tax_amount= 0.0
            if pur_line_id.taxes_id:

                for each_line in pur_line_id.taxes_id:
                    if each_line.tax_group_id.name == 'IGST':
                        for each_tax in each_line:
                            igst_rate  = each_tax.amount if each_tax.amount else 0.0
                    elif each_line.tax_group_id.name != 'IGST' and each_line.tax_group_id.name != 'GST':
                        for each_tax in each_line:
                            other_rate  = each_tax.amount if each_tax.amount else 0.0
                    elif each_line.tax_group_id.name == 'GST':
                        for each_tax in each_line.children_tax_ids:
                            sgst_rate  = each_tax.amount if each_tax.amount else 0.0
                            cgst_rate  = each_tax.amount if each_tax.amount else 0.0
            sl_val = ws.cell(row=row_c, column=1, value=sl_num)
            sl_val.alignment = Alignment(horizontal='center',vertical='center')
            item_category = ws.cell(row=row_c, column=2, value=pur_line_id.product_id.categ_id.complete_name)


            po_create_date = ws.cell(row=row_c, column=3, value=po_line['create_date'])
            po_create_date .alignment = Alignment(horizontal='center',vertical='center')

            po_m_date = ws.cell(row=row_c, column=4, value=po_line['write_date'])
            po_m_date.alignment = Alignment(horizontal='center',vertical='center')

            po_date = ws.cell(row=row_c, column=5, value=po_line['date_approve'])
            po_date.alignment = Alignment(horizontal='center',vertical='center')

            
            po_deli_date = ws.cell(row=row_c, column=6, value=po_line['de_date'])
            po_deli_date .alignment = Alignment(horizontal='center',vertical='center')
            if po_line['create_uid']:
                c_name = self.env['res.users'].search([('id','=',po_line['create_uid'])])

            
            created_by = ws.cell(row=row_c, column=7, value=c_name.name)
            created_by.alignment = Alignment(horizontal='center',vertical='center')

            supplier_name = ws.cell(row=row_c, column=8, value=pur_line_id.order_id.partner_id.name)
            supplier_name .alignment = Alignment(horizontal='center',vertical='center')


            warehouse_name = ws.cell(row=row_c, column=9, value= po_line['w_name'])
            # po_type = ws.cell(row=row_c, column=11, value=pur_line_id.order_id.name)
            po_state = ws.cell(row=row_c, column=10, value=po_line['state'])
            po_name = ws.cell(row=row_c, column=11, value=pur_line_id.order_id.name)
            # # # print("current_prod['lot_name']current_prod['lot_name']current_prod['lot_name']",current_prod['lot_name'])
            product_code = ws.cell(row=row_c, column=12, value=pur_line_id.product_id.default_code )
            product_id = ws.cell(row=row_c, column=13, value=pur_line_id.product_id.name )
            # pack_size = ws.cell(row=row_c, column=16, value=pur_line_id.product_uom.name )
            po_qty = ws.cell(row=row_c, column=14, value=pur_line_id.product_qty)
            po_uom = ws.cell(row=row_c, column=15, value=pur_line_id.product_uom.name )
            
            po_price = ws.cell(row=row_c, column=16, value=pur_line_id.price_unit)
            # discount = ws.cell(row=row_c, column=20, value=grn_line['grn_price'])
            if pur_line_id.price_subtotal :
                g_v =pur_line_id.price_subtotal
                cgst_amount =(g_v *cgst_rate )/100
                sgst_amount =(g_v *cgst_rate )/100
                igst_amount =(g_v *igst_rate )/100
                other_amount =(g_v *other_rate )/100
                total_tax_amount = cgst_amount+igst_amount+sgst_amount+other_amount
            else:
                g_v =0.0
            po_value = ws.cell(row=row_c, column=17, value=pur_line_id.price_subtotal)
            # gst = ws.cell(row=row_c, column=22, value=cgst_rate + sgst_rate)
            # gst = ws.cell(row=row_c, column=23, value=(sgst_amount+cgst_amount))
            cgst = ws.cell(row=row_c, column=18, value=cgst_rate)
            cgst_amt = ws.cell(row=row_c, column=19, value=(cgst_amount))
            s_gst = ws.cell(row=row_c, column=20, value=sgst_rate)
            s_gst_amt = ws.cell(row=row_c, column=21, value=(sgst_amount))
            igst = ws.cell(row=row_c, column=22, value=igst_rate)
            igst_amt = ws.cell(row=row_c, column=23, value=(igst_amount))


            other_tax = ws.cell(row=row_c, column=24, value=other_rate)
            other_tax_amt = ws.cell(row=row_c, column=25, value=(other_amount))
            po_tot = ws.cell(row=row_c, column=26, value=g_v+ total_tax_amount)

            default_uom = ws.cell(row=row_c, column=27, value=move_id.quantity_done)
            qty_uom = ws.cell(row=row_c, column=28, value= pur_line_id.product_id.uom_id.name)
            if pur_line_id.product_uom.uom_type == 'bigger':
                uom_rate = pur_line_id.price_unit/pur_line_id.product_uom.factor_inv
            elif  pur_line_id.product_uom.uom_type == 'smaller' :
                uom_rate = pur_line_id.price_unit*pur_line_id.product_uom.factor_inv
            else:
                uom_rate =pur_line_id.price_unit

            rate_uom = ws.cell(row=row_c, column=29, value=uom_rate)

            sl_num +=1
            row_c +=1



       




        fp = io.BytesIO()
        wb.save(fp)
        excel_file = base64.encodestring(fp.getvalue())
        self.po_report = excel_file
        self.po_printed = True
        self.file_name = "Purchase.xlsx"
        fp.close()

        return {
        'view_mode': 'form',
        'res_id': self.id,
        'res_model': 'purchase.report.excel',
        'view_type': 'form',
        'type': 'ir.actions.act_window',
        'context': self.env.context,
        'target': 'new',
                   }


   
