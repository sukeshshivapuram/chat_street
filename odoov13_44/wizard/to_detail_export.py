# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.



import io
import locale
import base64
import textwrap
from copy import copy

from datetime import datetime
from openpyxl import Workbook
from odoo import models, fields, api,_
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from odoo.exceptions import UserError, ValidationError,Warning
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
import pdb

class ToDetailExport(models.TransientModel):
    _name = "to.detail.export"
    _description = "to.detail.export"

    date_start = fields.Datetime(string="Start Date", required=True)
    date_end = fields.Datetime(string="End Date", required=True, default=fields.Datetime.now)
    to_report = fields.Binary('To Detail Export Report')
    file_name = fields.Char('File Name')
    to_printed = fields.Boolean('To Detail Export Printed')
    
    

    @api.constrains('date_start')
    def _code_constrains(self):
        if self.date_start > self.date_end:
            raise ValidationError(_("'Start Date' must be before 'End Date'"))


    

    def generate_report(self):

    
        # #Create Workbook and Worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "To-Detail Export Report"
        report_heading = " To Detail Export Report from" + ' ' + datetime.strftime(self.date_start, '%d-%m-%Y') + ' '+ 'To' + ' '+ datetime.strftime( self.date_end, '%d-%m-%Y')
        
        # #Border
        thin = Side(border_style="thin", color="000000")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=18)

        comapny = ws.cell(row=1, column=1, value=self.env.user.company_id.name + report_heading)
        comapny.alignment = Alignment(horizontal='center',vertical='center')
        comapny.border = Border(bottom=thin,top=thin)
        comapny.font = Font(size=10,name='Calibri')


        # Defining the Table Caolumn headings
        sl_no = ws.cell(row=3, column=1, value="S.No")
        sl_no.alignment = copy(comapny.alignment)

        # date = ws.cell(row=3, column=2, value="Date")
        # date.alignment = copy(comapny.alignment)

        sto_date = ws.cell(row=3, column=2, value="Date")
        sto_date.alignment = copy(comapny.alignment)

        dc_num = ws.cell(row=3, column=3, value="TO Number")
        dc_num.alignment = Alignment(horizontal='center',vertical='center')

        dc_state = ws.cell(row=3, column=4, value="TO Status")
        dc_state.alignment = Alignment(horizontal='center',vertical='center')

        sto_name = ws.cell(row=3, column=5, value="PO Number")
        sto_name.alignment = Alignment(horizontal='center',vertical='center')

        supply_site = ws.cell(row=3, column=6, value="Supply to Site")
        supply_site.alignment = Alignment(horizontal='center',vertical='center')

        supply_name = ws.cell(row=3, column=7, value="Supplier Name")
        supply_name.alignment = Alignment(horizontal='center',vertical='center')

        product_cat = ws.cell(row=3, column=8, value="Item Category Name")
        product_cat.alignment = Alignment(horizontal='center',vertical='center')

        item_name = ws.cell(row=3, column=9, value="Item Name")
        item_name.alignment = Alignment(horizontal='center',vertical='center')
        
        p_unit = ws.cell(row=3, column=10, value="Unit")
        p_unit.alignment = Alignment(horizontal='center',vertical='center')


        to_qty = ws.cell(row=3, column=11, value="TO Qty")
        to_qty.alignment = Alignment(horizontal='center',vertical='center')

        to_rate = ws.cell(row=3, column=12, value="Rate")
        to_rate.alignment = Alignment(horizontal='center',vertical='center')

        to_amt = ws.cell(row=3, column=13, value="TO Amt")
        to_amt.alignment = Alignment(horizontal='center',vertical='center')

        
        
        

        # query_sto = '''
        # select  
        #     ict.id, 
        #     ict.name as ict_name,
        #     ict.processed_date as ict_date,
        #     sp.id,
        #     sp.name 
        # from    
        #     stock_picking sp,
        #     inter_company_transfer_ept as ict 
        # where
        
        #     sp.inter_company_transfer_id =ict.id  and
        #     ict.processed_date > %s  and 
        #     ict.processed_date < %s and
        #     ict.state = 'processed' 

        # '''

        # params = (self.date_start ,self.date_end)

        # self.env.cr.execute(query_sto, params)
        # result = self.env.cr.dictfetchall()


        
        row_c=4
        sl_num=1

        # for grn_line in result:
        EPT = self.env['inter.company.transfer.ept']
        StockPicking = self.env['stock.picking']
        picking_rec='''

        select  
                sp.id as dc_id,
                sp.name as dc_name,
                sp.state as dc_state,
                ict.name as sto_name,
                ict.source_warehouse_id as sw_id,
                ict.destination_warehouse_id as dw_id,
                ict.processed_date as ict_date
        from  
                stock_picking as sp,
                inter_company_transfer_ept ict 

        where   
            sp.scheduled_date >= %s and 
            sp.scheduled_date <= %s and 
            sp.inter_company_transfer_id  = ict.id and 
            sp.picking_type_name = 'Delivery Orders'
                '''
        params = (self.date_start,self.date_end)
        self.env.cr.execute(picking_rec, params)
        picking_result = self.env.cr.dictfetchall()
        

        for each_picking in picking_result:

            # picking =StockPicking.search([('processed_date', '>',  self.date_start) , ('processed_date', '<', self.date_end)])
            if each_picking:
                # Ge the GRN product quantity group by product
                query_move_line ='''
                select  *

                from    
                stock_move_line 
                where 
                picking_id = %s 
                '''
                params = (each_picking['dc_id'],)
                # 
                self.env.cr.execute(query_move_line, params)
                result_line = self.env.cr.dictfetchall()
                if not result_line:
                    query_move ='''
                    select  *

                    from    
                    stock_move 
                    where 
                    picking_id = %s 
                    '''
                    params = (each_picking['dc_id'],)
                    # 
                    self.env.cr.execute(query_move, params)
                    result_move = self.env.cr.dictfetchall()
                if result_line:
                    result=result_line
                else:
                    result=result_move
                    # pdb.set_trace()


                for each_product in result:
                    p_quer='''
                    select 
                        pt.name as p_name,
                        uu.name as uu_name,
                        pc.complete_name  as pc_name
                    from  
                        product_product as pp,
                        product_template as pt,
                        uom_uom as uu,
                        product_category as pc
                    where 
                        product_tmpl_id = pt.id and 
                        pp.id =%s and
                        uu.id = pt.uom_id and
                        pc.id = pt.categ_id
                    '''
                    self.env.cr.execute(p_quer, [each_product['product_id']])
                    product = self.env.cr.dictfetchall()




                    # Get the warehouse names 
                    sw_query='''
                    select  
                        *
                    from 
                        stock_warehouse 
                    where 
                        id = %s 
                        '''
                    params_sw=[each_picking['sw_id']]
                    params_dw=[each_picking['dw_id']]

                    self.env.cr.execute(sw_query,params_sw )
                    sw_result = self.env.cr.dictfetchall()

                    self.env.cr.execute(sw_query,params_dw )
                    dw_result = self.env.cr.dictfetchall()


                    sl_val = ws.cell(row=row_c, column=1, value=sl_num)
                    sl_val.alignment = Alignment(horizontal='center',vertical='center')


                    sto_date = ws.cell(row=row_c, column=2, value=each_picking['ict_date'] )
                    sto_date .alignment = Alignment(horizontal='center',vertical='center')

                    dc_num = ws.cell(row=row_c, column=3, value=each_picking['dc_name'])
                    dc_num .alignment = Alignment(horizontal='center',vertical='center')
                    if each_picking['dc_state'] =='done':
                        state= 'Done'
                    elif each_picking['dc_state'] =='draft':
                        state= 'Draft'
                    elif each_picking['dc_state'] =='waiting':
                        state= 'Waiting Another Operation'
                    elif each_picking['dc_state'] =='confirmed':
                        state= 'Waiting'
                    elif each_picking['dc_state'] =='assigned':
                        state= 'Ready'
                    elif each_picking['dc_state'] =='cancel':
                        state= 'Cancelled'

                    dc_state = ws.cell(row=row_c, column=4, value=state)
                    # dc_state.alignment = Alignment(horizontal='center',vertical='center')

                    sto_name = ws.cell(row=row_c, column=5, value=each_picking['sto_name'])
                    # sto_name.alignment = Alignment(horizontal='center',vertical='center')

                    supply_site = ws.cell(row=row_c, column=6, value=dw_result[0]['name'])
                    # supply_site.alignment = Alignment(horizontal='center',vertical='center')


                    supply_name = ws.cell(row=row_c, column=7, value=sw_result[0]['name'])
                    # supply_name .alignment = Alignment(horizontal='center',vertical='center')


                    product_cat = ws.cell(row=row_c, column=8, value=product[0]['pc_name'])


                    item_name = ws.cell(row=row_c, column=9, value=product[0]['p_name'])



                    p_unit = ws.cell(row=row_c, column=10, value= product[0]['uu_name'])
                    if each_picking['dc_state'] =='done':
                        report_qty= each_product['qty_done']
                    else:
                        report_qty= each_product['product_uom_qty']
                    price_unit=0.0
                    curr_prod = self.env['product.product'].search([('id','=',each_product['product_id'])])
                    if curr_prod:
                        tax_per= 0.0
                        for each_tax in curr_prod.supplier_taxes_id:
                            if each_tax.tax_group_id.name == "GST" :
                                for each_gst in each_tax.children_tax_ids:
                                    tax_per+= each_gst.amount
                            elif  each_tax.tax_group_id.name == "IGST":
                                tax_per+= each_tax.amount
                            else:
                                tax_per+= each_tax.amount
                    # print("tax_pertax_pertax_per",tax_per)

                    move_line_id = self.env['stock.move.line'].search([('id','=',each_product['id'])])

                    if move_line_id.move_id.z_price>0:
                        tax_amt = move_line_id.move_id.z_price*(tax_per/100)
                        price_unit= move_line_id.move_id.z_price+ tax_amt
                    else:
                        tax_amt = curr_prod.standard_price*(tax_per/100)
                        price_unit =curr_prod.standard_price + tax_amt


                    to_qty = ws.cell(row=row_c, column=11, value=report_qty)
                    to_rate = ws.cell(row=row_c, column=12, value=price_unit)

                    to_amt = ws.cell(row=row_c, column=13, value=report_qty*price_unit)


                    #         # gr_qty = ws.cell(row=row_c, column=13, value=each_sm['grn_qty'] )

                    #         # grn_rate = ws.cell(row=row_c, column=14, value= curr_prod.standard_price)

                    #         # gr_amt = ws.cell(row=row_c, column=15, value=curr_prod.standard_price* each_sm['grn_qty'])

                    #         # diff_qty = ws.cell(row=row_c, column=16, value=sto_result[0]['s_qty']-each_sm['grn_qty'])

                    #         # diff_amt = ws.cell(row=row_c, column=17, value=(sto_result[0]['s_qty']*curr_prod.standard_price)-(curr_prod.standard_price* each_sm['grn_qty']))

                    #         # city = ws.cell(row=row_c, column=18, value= "True" if each_ict.destination_warehouse_id.city_id.id == each_ict.source_warehouse_id.city_id.id else "False" )



                    sl_num +=1
                    row_c +=1



        


        fp = io.BytesIO()
        wb.save(fp)
        excel_file = base64.encodestring(fp.getvalue())
        self.to_report = excel_file
        self.to_printed = True
        self.file_name = "To Detail Export.xlsx"
        fp.close()

        return {
        'view_mode': 'form',
        'res_id': self.id,
        'res_model': 'to.detail.export',
        'view_type': 'form',
        'type': 'ir.actions.act_window',
        'context': self.env.context,
        'target': 'new',
                   }


   
